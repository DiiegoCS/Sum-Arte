"""
Módulo de generación de reportes para Sum-Arte.

Este módulo contiene funciones para generar reportes en PDF y Excel,
incluyendo reportes de estado de proyectos y reportes de rendición oficial.
"""

from io import BytesIO
from decimal import Decimal
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from .models import (
    Proyecto, Transaccion, Item_Presupuestario, Subitem_Presupuestario,
    Evidencia, Transaccion_Evidencia, Organizacion
)


def generar_reporte_estado_proyecto_pdf(proyecto):
    """
    Genera un reporte PDF del estado de un proyecto activo.
    
    El reporte incluye:
    - Resumen ejecutivo (presupuesto, ejecutado, disponible)
    - Lista de transacciones aprobadas con evidencias
    - Ítems presupuestarios con ejecución
    - Gráficos de ejecución (texto descriptivo)
    - Estado de rendición
    
    Args:
        proyecto: Instancia de Proyecto
        
    Returns:
        BytesIO: Buffer con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    # Contenedor para los elementos del documento
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Título
    elements.append(Paragraph("REPORTE DE ESTADO DE PROYECTO", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información del proyecto
    info_data = [
        ['Proyecto:', proyecto.nombre_proyecto],
        ['Organización:', proyecto.id_organizacion.nombre_organizacion],
        ['Estado:', proyecto.get_estado_proyecto_display() if hasattr(proyecto, 'get_estado_proyecto_display') else proyecto.estado_proyecto],
        ['Fecha de inicio:', proyecto.fecha_inicio_proyecto.strftime('%d/%m/%Y')],
        ['Fecha de fin:', proyecto.fecha_fin_proyecto.strftime('%d/%m/%Y')],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumen ejecutivo
    elements.append(Paragraph("RESUMEN EJECUTIVO", heading_style))
    
    presupuesto_total = float(proyecto.presupuesto_total)
    monto_ejecutado = float(proyecto.monto_ejecutado_proyecto)
    monto_disponible = presupuesto_total - monto_ejecutado
    porcentaje_ejecutado = (monto_ejecutado / presupuesto_total * 100) if presupuesto_total > 0 else 0
    
    resumen_data = [
        ['Concepto', 'Monto (CLP)'],
        ['Presupuesto Total', f'${presupuesto_total:,.0f}'],
        ['Monto Ejecutado', f'${monto_ejecutado:,.0f}'],
        ['Monto Disponible', f'${monto_disponible:,.0f}'],
        ['Porcentaje Ejecutado', f'{porcentaje_ejecutado:.2f}%'],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 3*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Transacciones aprobadas
    transacciones = Transaccion.objects.filter(
        proyecto=proyecto,
        estado_transaccion='aprobado'
    ).select_related('proveedor', 'usuario', 'item_presupuestario', 'subitem_presupuestario').order_by('-fecha_registro')
    
    if transacciones.exists():
        elements.append(Paragraph("TRANSACCIONES APROBADAS", heading_style))
        
        # Encabezados de tabla
        trans_data = [['Fecha', 'Proveedor', 'Documento', 'Monto', 'Evidencias']]
        
        for trans in transacciones[:50]:  # Limitar a 50 para no sobrecargar el PDF
            # Contar evidencias
            evidencias_count = Transaccion_Evidencia.objects.filter(
                transaccion=trans,
                evidencia__eliminado=False
            ).count()
            
            trans_data.append([
                trans.fecha_registro.strftime('%d/%m/%Y'),
                trans.proveedor.nombre_proveedor if trans.proveedor else 'N/A',
                trans.nro_documento or 'N/A',
                f'${float(trans.monto_transaccion):,.0f}',
                str(evidencias_count)
            ])
        
        trans_table = Table(trans_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1*inch, 0.8*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Monto alineado a la derecha
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(trans_table)
        
        if transacciones.count() > 50:
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph(
                f"<i>Nota: Se muestran las primeras 50 transacciones de un total de {transacciones.count()}</i>",
                styles['Normal']
            ))
        
        elements.append(Spacer(1, 0.3*inch))
    
    # Ítems presupuestarios
    items = Item_Presupuestario.objects.filter(proyecto=proyecto).prefetch_related('subitem_presupuestario_set')
    
    if items.exists():
        elements.append(Paragraph("ÍTEMS PRESUPUESTARIOS", heading_style))
        
        items_data = [['Ítem', 'Asignado', 'Ejecutado', 'Disponible', '% Ejecutado']]
        
        for item in items:
            asignado = float(item.monto_asignado_item)
            ejecutado = float(item.monto_ejecutado_item)
            disponible = asignado - ejecutado
            porcentaje = (ejecutado / asignado * 100) if asignado > 0 else 0
            
            items_data.append([
                item.nombre_item_presupuesto,
                f'${asignado:,.0f}',
                f'${ejecutado:,.0f}',
                f'${disponible:,.0f}',
                f'{porcentaje:.2f}%'
            ])
        
        items_table = Table(items_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),  # Montos alineados a la derecha
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Pie de página
    fecha_generacion = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(
        f"<i>Reporte generado el {fecha_generacion}</i>",
        styles['Normal']
    ))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_reporte_estado_proyecto_excel(proyecto):
    """
    Genera un reporte Excel del estado de un proyecto activo.
    
    Args:
        proyecto: Instancia de Proyecto
        
    Returns:
        BytesIO: Buffer con el archivo Excel generado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado del Proyecto"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"REPORTE DE ESTADO DE PROYECTO: {proyecto.nombre_proyecto}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 25
    
    # Información del proyecto
    row = 3
    ws[f'A{row}'] = 'Proyecto:'
    ws[f'B{row}'] = proyecto.nombre_proyecto
    row += 1
    ws[f'A{row}'] = 'Organización:'
    ws[f'B{row}'] = proyecto.id_organizacion.nombre_organizacion
    row += 1
    ws[f'A{row}'] = 'Estado:'
    ws[f'B{row}'] = proyecto.get_estado_proyecto_display() if hasattr(proyecto, 'get_estado_proyecto_display') else proyecto.estado_proyecto
    row += 1
    ws[f'A{row}'] = 'Fecha de inicio:'
    ws[f'B{row}'] = proyecto.fecha_inicio_proyecto.strftime('%d/%m/%Y')
    row += 1
    ws[f'A{row}'] = 'Fecha de fin:'
    ws[f'B{row}'] = proyecto.fecha_fin_proyecto.strftime('%d/%m/%Y')
    row += 2
    
    # Resumen ejecutivo
    ws[f'A{row}'] = 'RESUMEN EJECUTIVO'
    ws[f'A{row}'].font = title_font
    row += 1
    
    presupuesto_total = float(proyecto.presupuesto_total)
    monto_ejecutado = float(proyecto.monto_ejecutado_proyecto)
    monto_disponible = presupuesto_total - monto_ejecutado
    porcentaje_ejecutado = (monto_ejecutado / presupuesto_total * 100) if presupuesto_total > 0 else 0
    
    resumen_headers = ['Concepto', 'Monto (CLP)']
    for col, header in enumerate(resumen_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    row += 1
    resumen_data = [
        ['Presupuesto Total', presupuesto_total],
        ['Monto Ejecutado', monto_ejecutado],
        ['Monto Disponible', monto_disponible],
        ['Porcentaje Ejecutado', porcentaje_ejecutado],
    ]
    
    for concepto, monto in resumen_data:
        ws.cell(row=row, column=1).value = concepto
        ws.cell(row=row, column=1).border = border
        if isinstance(monto, float) and concepto == 'Porcentaje Ejecutado':
            ws.cell(row=row, column=2).value = f"{monto:.2f}%"
        else:
            ws.cell(row=row, column=2).value = f"${monto:,.0f}"
        ws.cell(row=row, column=2).alignment = right_align
        ws.cell(row=row, column=2).border = border
        row += 1
    
    row += 1
    
    # Transacciones aprobadas
    transacciones = Transaccion.objects.filter(
        proyecto=proyecto,
        estado_transaccion='aprobado'
    ).select_related('proveedor', 'item_presupuestario', 'subitem_presupuestario').order_by('-fecha_registro')
    
    if transacciones.exists():
        ws[f'A{row}'] = 'TRANSACCIONES APROBADAS'
        ws[f'A{row}'].font = title_font
        row += 1
        
        trans_headers = ['Fecha', 'Proveedor', 'Documento', 'Monto', 'Evidencias']
        for col, header in enumerate(trans_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        
        for trans in transacciones:
            evidencias_count = Transaccion_Evidencia.objects.filter(
                transaccion=trans,
                evidencia__eliminado=False
            ).count()
            
            ws.cell(row=row, column=1).value = trans.fecha_registro.strftime('%d/%m/%Y')
            ws.cell(row=row, column=2).value = trans.proveedor.nombre_proveedor if trans.proveedor else 'N/A'
            ws.cell(row=row, column=3).value = trans.nro_documento or 'N/A'
            ws.cell(row=row, column=4).value = f"${float(trans.monto_transaccion):,.0f}"
            ws.cell(row=row, column=5).value = evidencias_count
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                if col == 4:
                    ws.cell(row=row, column=col).alignment = right_align
            
            row += 1
        
        row += 1
    
    # Ítems presupuestarios
    items = Item_Presupuestario.objects.filter(proyecto=proyecto)
    
    if items.exists():
        ws[f'A{row}'] = 'ÍTEMS PRESUPUESTARIOS'
        ws[f'A{row}'].font = title_font
        row += 1
        
        items_headers = ['Ítem', 'Asignado', 'Ejecutado', 'Disponible', '% Ejecutado']
        for col, header in enumerate(items_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        
        for item in items:
            asignado = float(item.monto_asignado_item)
            ejecutado = float(item.monto_ejecutado_item)
            disponible = asignado - ejecutado
            porcentaje = (ejecutado / asignado * 100) if asignado > 0 else 0
            
            ws.cell(row=row, column=1).value = item.nombre_item_presupuesto
            ws.cell(row=row, column=2).value = f"${asignado:,.0f}"
            ws.cell(row=row, column=3).value = f"${ejecutado:,.0f}"
            ws.cell(row=row, column=4).value = f"${disponible:,.0f}"
            ws.cell(row=row, column=5).value = f"{porcentaje:.2f}%"
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                if col in [2, 3, 4]:
                    ws.cell(row=row, column=col).alignment = right_align
            
            row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_reporte_rendicion_oficial_pdf(proyecto):
    """
    Genera el reporte oficial de rendición en PDF.
    
    Este reporte es más formal y completo que el de estado, ya que es el documento
    oficial que se presenta al cierre de la rendición. Incluye:
    - Portada con información completa
    - Resumen ejecutivo detallado
    - Lista completa de transacciones aprobadas con evidencias
    - Desglose por ítems presupuestarios
    - Certificación de cierre
    
    Args:
        proyecto: Instancia de Proyecto (debe estar cerrado/completado)
        
    Returns:
        BytesIO: Buffer con el PDF generado
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'OfficialTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=40,
        alignment=1,  # Centrado
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'OfficialSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=1,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'OfficialHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=20,
        fontName='Helvetica-Bold'
    )
    
    # PORTADA
    elements.append(Spacer(1, 2*inch))
    elements.append(Paragraph("REPORTE OFICIAL DE RENDICIÓN", title_style))
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph(proyecto.nombre_proyecto, subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(
        f"<b>Organización:</b> {proyecto.id_organizacion.nombre_organizacion}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph(
        f"<b>RUT:</b> {proyecto.id_organizacion.rut_organizacion}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph(
        f"<b>Período:</b> {proyecto.fecha_inicio_proyecto.strftime('%d/%m/%Y')} - {proyecto.fecha_fin_proyecto.strftime('%d/%m/%Y')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.1*inch))
    fecha_cierre = timezone.now()
    elements.append(Paragraph(
        f"<b>Fecha de cierre:</b> {fecha_cierre.strftime('%d/%m/%Y %H:%M:%S')}",
        styles['Normal']
    ))
    elements.append(PageBreak())
    
    # INFORMACIÓN DEL PROYECTO
    elements.append(Paragraph("INFORMACIÓN DEL PROYECTO", heading_style))
    
    info_data = [
        ['Campo', 'Valor'],
        ['Nombre del Proyecto', proyecto.nombre_proyecto],
        ['Organización', proyecto.id_organizacion.nombre_organizacion],
        ['RUT Organización', proyecto.id_organizacion.rut_organizacion],
        ['Estado', proyecto.get_estado_proyecto_display() if hasattr(proyecto, 'get_estado_proyecto_display') else proyecto.estado_proyecto],
        ['Fecha de Inicio', proyecto.fecha_inicio_proyecto.strftime('%d/%m/%Y')],
        ['Fecha de Fin', proyecto.fecha_fin_proyecto.strftime('%d/%m/%Y')],
        ['Fecha de Cierre de Rendición', fecha_cierre.strftime('%d/%m/%Y %H:%M:%S')],
    ]
    
    info_table = Table(info_data, colWidths=[2.5*inch, 3.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # RESUMEN EJECUTIVO
    elements.append(Paragraph("RESUMEN EJECUTIVO", heading_style))
    
    presupuesto_total = float(proyecto.presupuesto_total)
    monto_ejecutado = float(proyecto.monto_ejecutado_proyecto)
    monto_disponible = presupuesto_total - monto_ejecutado
    porcentaje_ejecutado = (monto_ejecutado / presupuesto_total * 100) if presupuesto_total > 0 else 0
    
    resumen_data = [
        ['Concepto', 'Monto (CLP)', 'Porcentaje'],
        ['Presupuesto Total Asignado', f'${presupuesto_total:,.0f}', '100.00%'],
        ['Monto Ejecutado', f'${monto_ejecutado:,.0f}', f'{porcentaje_ejecutado:.2f}%'],
        ['Monto Disponible', f'${monto_disponible:,.0f}', f'{(100 - porcentaje_ejecutado):.2f}%'],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # TRANSACCIONES APROBADAS
    transacciones = Transaccion.objects.filter(
        proyecto=proyecto,
        estado_transaccion='aprobado'
    ).select_related(
        'proveedor', 'usuario', 'item_presupuestario', 'subitem_presupuestario'
    ).order_by('fecha_registro')
    
    if transacciones.exists():
        elements.append(Paragraph("DETALLE DE TRANSACCIONES APROBADAS", heading_style))
        
        trans_data = [['#', 'Fecha', 'Proveedor', 'Documento', 'Monto', 'Evidencias']]
        
        total_transacciones = 0
        for idx, trans in enumerate(transacciones, 1):
            evidencias_count = Transaccion_Evidencia.objects.filter(
                transaccion=trans,
                evidencia__eliminado=False
            ).count()
            
            total_transacciones += float(trans.monto_transaccion)
            
            trans_data.append([
                str(idx),
                trans.fecha_registro.strftime('%d/%m/%Y'),
                (trans.proveedor.nombre_proveedor[:30] + '...') if trans.proveedor and len(trans.proveedor.nombre_proveedor) > 30 else (trans.proveedor.nombre_proveedor if trans.proveedor else 'N/A'),
                trans.nro_documento or 'N/A',
                f'${float(trans.monto_transaccion):,.0f}',
                str(evidencias_count)
            ])
        
        # Fila de total
        trans_data.append([
            '',
            '',
            '<b>TOTAL</b>',
            '',
            f'<b>${total_transacciones:,.0f}</b>',
            ''
        ])
        
        trans_table = Table(trans_data, colWidths=[0.4*inch, 0.9*inch, 1.3*inch, 0.9*inch, 1*inch, 0.5*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -2), 'LEFT'),
            ('ALIGN', (4, 1), (4, -2), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(trans_table)
        elements.append(Spacer(1, 0.4*inch))
    
    # ÍTEMS PRESUPUESTARIOS
    items = Item_Presupuestario.objects.filter(proyecto=proyecto).prefetch_related('subitem_presupuestario_set')
    
    if items.exists():
        elements.append(Paragraph("DESGLOSE POR ÍTEMS PRESUPUESTARIOS", heading_style))
        
        items_data = [['Ítem', 'Asignado', 'Ejecutado', 'Disponible', '% Ejecutado']]
        
        for item in items:
            asignado = float(item.monto_asignado_item)
            ejecutado = float(item.monto_ejecutado_item)
            disponible = asignado - ejecutado
            porcentaje = (ejecutado / asignado * 100) if asignado > 0 else 0
            
            items_data.append([
                item.nombre_item_presupuesto,
                f'${asignado:,.0f}',
                f'${ejecutado:,.0f}',
                f'${disponible:,.0f}',
                f'{porcentaje:.2f}%'
            ])
        
        items_table = Table(items_data, colWidths=[2.5*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 0.4*inch))
    
    # CERTIFICACIÓN
    elements.append(PageBreak())
    elements.append(Paragraph("CERTIFICACIÓN", heading_style))
    elements.append(Spacer(1, 0.3*inch))
    
    certificacion_texto = f"""
    <para>
    Por medio del presente documento, se certifica que la rendición del proyecto 
    <b>"{proyecto.nombre_proyecto}"</b> ha sido cerrada y validada según los controles 
    establecidos en el sistema Sum-Arte.
    </para>
    
    <para>
    Se confirma que:
    </para>
    
    <para>
    • Todas las transacciones han sido aprobadas y cuentan con evidencia documental vinculada.
    </para>
    
    <para>
    • El presupuesto ejecutado corresponde a la suma de las transacciones aprobadas.
    </para>
    
    <para>
    • La rendición cumple con los requisitos de integridad y trazabilidad establecidos.
    </para>
    
    <para>
    Este reporte fue generado automáticamente el {fecha_cierre.strftime('%d de %B de %Y a las %H:%M:%S')} 
    y constituye el documento oficial de rendición del proyecto.
    </para>
    """
    
    elements.append(Paragraph(certificacion_texto, styles['Normal']))
    elements.append(Spacer(1, 0.5*inch))
    
    # Pie de página
    fecha_generacion = fecha_cierre.strftime('%d/%m/%Y %H:%M:%S')
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(
        f"<i>Reporte oficial generado el {fecha_generacion}</i>",
        styles['Normal']
    ))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

